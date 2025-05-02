import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

def run_test(user, password):
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    driver = webdriver.Chrome(chrome_options)
    driver.get("http://demostore.supersqa.com/my-account/")
    driver.implicitly_wait(10)
    driver.maximize_window()

    username_fld = driver.find_element(By.ID, 'username')
    password_fld = driver.find_element(By.ID, 'password')

    # update password field attribute with js script to make its value visible
    driver.execute_script('arguments[0].setAttribute("type", "text")', password_fld)

    username_fld.send_keys(user)
    password_fld.send_keys(password)
    time.sleep(2)
    username_fld.clear()
    password_fld.clear()
    time.sleep(3)
    driver.quit()


# IDE offers to install
from openpyxl import load_workbook
# guide
# https://www.youtube.com/watch?v=_ZEzMMMytRs

# load sheet from xls file
# workbook = load_workbook('xls_demo.xlsx')
workbook = load_workbook('/Users/alexeyvlasov/Documents/Automation/Selemium-Python/Selenium/my_scripts/data_driven/tables/xls_demo.xlsx')
# select active sheet
sheet = workbook.active

# output for test data
# test_data = []

# set starting (min row), max row and ignore empty cells
for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, values_only = True):
    # pick first and second values from each row
    user = row[0]
    psw = row[1]
    # print('username: {}; password: {}'.format(user, psw))

    # run test
    run_test(user, psw)

    #each row is a tuple  - form list of tuples
    # test_data.append(row)

#printing test data
# print(test_data)